"""
Automação: lê números SGM de uma planilha no Excel Online, pesquisa cada um
no ServiceNow e clica no link/campo MAN do chamado correspondente.

LOGIN: este script assume que você já está (ou vai ficar) logado num perfil
de navegador persistente — não digita usuário/senha automaticamente.
Veja o README.md para o passo a passo completo.

  Primeiro uso (login manual):   python main.py --login
  Uso normal (já logado):        python main.py

TUDO QUE PRECISA SER AJUSTADO PARA O SEU AMBIENTE está marcado com:
  🔧 AJUSTE AQUI
Veja no README.md "Como encontrar o seletor certo" o passo a passo de como
descobrir o valor certo usando o Inspecionar do navegador (F12).
"""

import csv
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from playwright.sync_api import sync_playwright, Page, BrowserContext

load_dotenv()

EXCEL_ONLINE_URL = os.getenv("EXCEL_ONLINE_URL")
EXCEL_LOCAL_PATH = os.getenv("EXCEL_LOCAL_PATH", "").strip().strip('"').strip("'")
SERVICENOW_URL = os.getenv("SERVICENOW_URL")
SERVICENOW_TASK_LIST_URL = os.getenv("SERVICENOW_TASK_LIST_URL", SERVICENOW_URL)
DOWNLOAD_DIR = Path(os.getenv("DOWNLOAD_DIR", "./downloads")).resolve()
USER_DATA_DIR = Path(os.getenv("USER_DATA_DIR", "./perfil_navegador")).resolve()

DOWNLOAD_DIR.mkdir(exist_ok=True, parents=True)
USER_DATA_DIR.mkdir(exist_ok=True, parents=True)

ERROS_DIR = Path("./erros").resolve()
ERROS_DIR.mkdir(exist_ok=True, parents=True)

RESULTADOS_DIR = Path("./resultados").resolve()
RESULTADOS_DIR.mkdir(exist_ok=True, parents=True)

COLUNA_SGM_INDEX = 4                  # Coluna E => "Número Chamado SGM"
COLUNA_TECNICO_INDEX = 5              # Coluna F => nome do técnico (pra validar contra o ServiceNow)
COLUNA_VALIDADOR_INDEX = 6            # Coluna G => "Coordenador (a)" (nome do validador)
COLUNA_MATRICULA_INDEX = 7            # Coluna H => "Técnico" (matrícula do validador)
COLUNA_DATA_ATENDIMENTO_INDEX = 8     # Coluna I => "Data de atendimento do chamado"
COLUNA_HORARIO_INICIO_INDEX = 9       # Coluna J => "Horário de inicio do atendimento"
COLUNA_HORARIO_TERMINO_INDEX = 10     # Coluna K => "Horário do fim do atendimento"


# ============================================================================
# ETAPA 1 — Baixar a planilha atualizada do Excel Online
# ============================================================================
def baixar_planilha_excel_online(page: Page) -> Path:
    """
    Confirmado com teste real: esse link do SharePoint serve o arquivo como
    download direto ao ser acessado, em vez de abrir o editor web do Excel
    Online. Por isso, não precisamos simular cliques em "Arquivo > Salvar
    uma Cópia" — basta capturar o download que acontece ao navegar pra essa
    URL. O try/except é necessário porque o próprio page.goto() "falha" (do
    ponto de vista do Playwright) quando a navegação se transforma num
    download em vez de carregar uma página normal — isso é esperado.
    """
    with page.expect_download() as download_info:
        try:
            page.goto(EXCEL_ONLINE_URL)
        except Exception:
            pass  # esperado: a navegação "falha" porque virou download direto
    download = download_info.value

    caminho_arquivo = DOWNLOAD_DIR / "planilha_chamados.xlsx"
    download.save_as(caminho_arquivo)
    return caminho_arquivo


# ============================================================================
# ETAPA 2 — Ler os dados da planilha (SGM, início real, término real)
# ============================================================================
def combinar_data_hora(data_valor, hora_valor):
    """
    Combina uma data (coluna I) com um horário (coluna J ou K) num único
    valor datetime. Lida com os formatos mais comuns que o Excel/openpyxl
    podem entregar: datetime, date, time, ou texto.

    Retorna um objeto datetime se conseguir combinar, ou None se não
    conseguir interpretar algum dos dois valores (nesse caso, um aviso é
    impresso em ler_itens_da_planilha).
    """
    from datetime import datetime as dt, date as date_cls, time as time_cls

    # --- Extrai a parte da DATA ---
    if isinstance(data_valor, dt):
        data = data_valor.date()
    elif isinstance(data_valor, date_cls):
        data = data_valor
    elif isinstance(data_valor, str):
        data = None
        for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%d/%m/%y", "%m/%d/%y"):
            try:
                data = dt.strptime(data_valor.strip(), fmt).date()
                break
            except ValueError:
                continue
        if data is None:
            return None
    else:
        return None

    # --- Extrai a parte da HORA ---
    if isinstance(hora_valor, dt):
        hora = hora_valor.time()
    elif isinstance(hora_valor, time_cls):
        hora = hora_valor
    elif isinstance(hora_valor, str):
        hora = None
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                hora = dt.strptime(hora_valor.strip(), fmt).time()
                break
            except ValueError:
                continue
        if hora is None:
            return None
    else:
        return None

    return dt.combine(data, hora)


def formatar_data_hora_servicenow(valor) -> str:
    """
    Formata no padrão que o campo do ServiceNow realmente espera:
    DD/MM/AAAA HH:MM:SS — confirmado pelo próprio valor já existente no
    campo "wm_task.work_start" ("17/06/2026 21:33:14"). Isso é diferente
    do formato que a planilha mostra (estilo americano, "3/25/26..."), daí
    a necessidade de converter em vez de só copiar o texto.

    Se a célula vier como um valor real de data/hora (o caso mais comum,
    já que a coluna é formatada como data no Excel), formatamos
    diretamente. Se vier como texto puro por algum motivo, usamos como
    veio e avisamos — nesse caso vale checar manualmente se bateu certo.
    """
    if valor is None:
        return ""
    if hasattr(valor, "strftime"):
        return valor.strftime("%d/%m/%Y %H:%M:%S")
    texto = str(valor).strip()
    print(f"⚠️  Aviso: valor de data/hora veio como texto puro ('{texto}'), não consegui reformatar — confirme se está no formato certo (DD/MM/AAAA HH:MM:SS).")
    return texto


def normalizar_sgm(valor_bruto) -> str:
    """
    Normaliza o número SGM lido da planilha para o formato padrão
    "SGM0999999" (SGM seguido de 7 dígitos — o "0" do começo faz parte dos
    números). Trata dois casos:

    1. Célula com só os 7 números (ex: "0670464") -> vira "SGM0670464".
    2. Célula com "SGM" e um espaço antes dos números (ex: "SGM 0670464")
       -> remove o espaço, virando "SGM0670464".

    Qualquer espaço extra também é removido. Se o resultado não bater com o
    formato esperado (SGM + 7 dígitos), o valor é usado mesmo assim, mas um
    aviso é impresso pra você conferir manualmente.
    """
    texto = str(valor_bruto).strip()

    # Remove TODOS os espaços (cobre "SGM 0670464", "SGM0670464 ", etc.)
    texto = texto.replace(" ", "")

    # Se vier só com dígitos (sem o prefixo "SGM"), trata o caso clássico do
    # Excel "comer" o zero à esquerda: quando a célula é formatada como
    # número, "0751687" é lido como "751687" (perde o zero). Por isso
    # completamos com zeros à esquerda até ter 7 dígitos antes de adicionar
    # o "SGM" — assim "751687" vira "0751687" -> "SGM0751687".
    if texto.isdigit():
        texto = "SGM" + texto.zfill(7)

    # Conferência final: o padrão esperado é SGM + 7 dígitos.
    if not re.fullmatch(r"SGM\d{7}", texto):
        print(f"⚠️  Aviso: SGM '{texto}' (original: '{valor_bruto}') não está no "
              "formato esperado SGM0999999 (SGM + 7 dígitos) — confira manualmente.")

    return texto


def ler_itens_da_planilha(caminho_arquivo: Path) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)

    # 🔧 AJUSTE AQUI (se necessário) — Nome da aba
    # wb.active pega a aba que estava selecionada quando o arquivo foi salvo.
    # Se sua planilha de chamados estiver numa aba específica, troque para:
    #   ws = wb["NomeDaAba"]
    ws = wb.active

    # Confirmado com uma amostra real da planilha: estes são os cabeçalhos
    # esperados nas colunas E, J e K. Essa checagem avisa se a estrutura
    # mudar (ex: alguém inseriu uma coluna nova, deslocando tudo).
    colunas_esperadas = {
        COLUNA_SGM_INDEX: "Número Chamado SGM",
        # 🔧 AJUSTE AQUI, SE NECESSÁRIO: confirme se os cabeçalhos das colunas
        # F, G e H batem com estes. Se o cabeçalho não bater, o aviso vai
        # aparecer no terminal.
        COLUNA_TECNICO_INDEX: "Técnico",
        COLUNA_VALIDADOR_INDEX: "Coordenador (a)",
        COLUNA_MATRICULA_INDEX: "Técnico",
        COLUNA_DATA_ATENDIMENTO_INDEX: "Data de atendimento do chamado",
        COLUNA_HORARIO_INICIO_INDEX: "Horário de inicio do atendimento",
        COLUNA_HORARIO_TERMINO_INDEX: "Horário do fim do atendimento",
    }
    for indice, cabecalho_esperado in colunas_esperadas.items():
        cabecalho_real = ws.cell(row=1, column=indice + 1).value  # openpyxl é 1-based
        if cabecalho_real != cabecalho_esperado:
            print(
                f"⚠️  Aviso: esperava o cabeçalho '{cabecalho_esperado}' na coluna "
                f"{indice + 1}, mas encontrei '{cabecalho_real}'. A planilha pode "
                "ter mudado de estrutura."
            )

    itens = []
    # min_row=2 pula a linha de cabeçalho.
    for row in ws.iter_rows(min_row=2):
        celula_sgm = row[COLUNA_SGM_INDEX]
        if not celula_sgm.value:
            continue
        tecnico = row[COLUNA_TECNICO_INDEX].value
        validador = row[COLUNA_VALIDADOR_INDEX].value
        matricula = row[COLUNA_MATRICULA_INDEX].value

        data_atendimento = row[COLUNA_DATA_ATENDIMENTO_INDEX].value
        hora_inicio = row[COLUNA_HORARIO_INICIO_INDEX].value
        hora_termino = row[COLUNA_HORARIO_TERMINO_INDEX].value

        sgm_atual = normalizar_sgm(celula_sgm.value)

        inicio_combinado = combinar_data_hora(data_atendimento, hora_inicio)
        if inicio_combinado is None:
            print(f"⚠️  Aviso: não consegui combinar data+hora de início pro SGM "
                  f"{sgm_atual} (data='{data_atendimento}', hora='{hora_inicio}').")

        termino_combinado = combinar_data_hora(data_atendimento, hora_termino)
        if termino_combinado is None:
            print(f"⚠️  Aviso: não consegui combinar data+hora de término pro SGM "
                  f"{sgm_atual} (data='{data_atendimento}', hora='{hora_termino}').")

        itens.append({
            "linha": celula_sgm.row,  # número da linha na planilha (pra escrever o status depois)
            "sgm": sgm_atual,
            "inicio_real": formatar_data_hora_servicenow(inicio_combinado) if inicio_combinado else "",
            "termino_real": formatar_data_hora_servicenow(termino_combinado) if termino_combinado else "",
            "tecnico": str(tecnico).strip() if tecnico is not None else "",
            "validador": str(validador).strip() if validador is not None else "",
            "matricula": str(matricula).strip() if matricula is not None else "",
        })
    return itens


# ============================================================================
# ETAPA 3 — Pesquisar o SGM no ServiceNow, clicar no MAN e preencher as etapas
# ============================================================================
def obter_alvo_servicenow(page: Page):
    """
    Retorna o objeto correto para interagir com a tela: o iframe 'gsft_main'
    (interface clássica) ou a própria página (interface nova/Next Experience).
    Veja README.md > "Caso especial do ServiceNow" para saber qual é o seu caso.
    """
    frame = page.frame(name="gsft_main")
    return frame if frame else page


def montar_url_busca_por_sgm(sgm: str) -> str:
    """
    Confirmado com teste real feito na sua instância: a busca usa o campo
    'parent.number' (o número do registro pai/relacionado ao Work Order
    Task) com o operador STARTSWITH — ou seja, o SGM é o número do registro
    de origem (provavelmente um Incident ou similar), e o Work Order Task
    (numerado como "MANxxxxxxx") é um registro relacionado a ele.
    """
    base = SERVICENOW_TASK_LIST_URL.split("?")[0]
    return f"{base}?sysparm_query=parent.numberSTARTSWITH{sgm}"


def buscar_chamado_e_clicar_man(page: Page, sgm: str):
    if "login" in page.url.lower():
        raise RuntimeError(
            "Sessão do ServiceNow expirou ou não está logada. "
            "Rode: python main.py --login"
        )

    page.goto(montar_url_busca_por_sgm(sgm))
    page.wait_for_load_state("networkidle")

    alvo = obter_alvo_servicenow(page)

    # MAN não é um campo separado — é o próprio número do registro na tabela
    # wm_task (toda tarefa dessa tabela é numerada como "MAN0010001" etc.).
    # Por isso buscamos qualquer link cujo texto comece com "MAN" seguido de
    # números, em vez de procurar um texto fixo — assim funciona mesmo se a
    # lista filtrada trouxer o registro em posições diferentes na tela.
    link_man = alvo.get_by_text(re.compile(r"^MAN\d+"))

    # Verificação rápida: se não encontrou nenhum resultado, desiste já,
    # em vez de esperar o tempo padrão (30s) tentando clicar em algo que
    # não existe. Isso deixa o "pular pro próximo chamado" bem mais rápido
    # quando o SGM não é encontrado no ServiceNow.
    if link_man.count() == 0:
        raise RuntimeError(f"SGM {sgm} não encontrado no ServiceNow.")

    # 🔧 AJUSTE AQUI, SE NECESSÁRIO: se a lista filtrada trouxer mais de um
    # resultado (ex: SGM associado a mais de uma tarefa), .first pega o
    # primeiro — confirme se isso é o comportamento certo pro seu caso, ou
    # se precisa adicionar lógica para escolher entre vários resultados.
    link_man.first.click(timeout=5000)
    alvo.wait_for_load_state("networkidle")

    print(f"Chamado {sgm}: entrou no registro MAN com sucesso.")

    # Atualiza (recarrega) a página depois de entrar no chamado.
    page.reload()
    page.wait_for_load_state("networkidle")
    fechar_popup_chamados_pausados(page)  # reforço manual, caso o popup apareça aqui
    # o frame é recriado após o reload, então buscamos de novo
    return obter_alvo_servicenow(page)


def obter_ou_criar_coluna_status(ws) -> int:
    """
    Procura uma coluna com cabeçalho "Status" na planilha; se não existir,
    cria uma nova coluna logo depois da última usada. Retorna o índice
    (1-based) dessa coluna.
    """
    for cell in ws[1]:  # primeira linha (cabeçalhos)
        if cell.value == "Status":
            return cell.column
    nova_coluna = ws.max_column + 1
    ws.cell(row=1, column=nova_coluna, value="Status")
    return nova_coluna


def fechar_popup_chamados_pausados(page: Page) -> None:
    """
    Fecha automaticamente o popup "Tem certeza? ... Chamado(s): MANxxxxxxx"
    (lista de chamados em andamento que seriam pausados) sempre que ele
    aparecer, clicando em "OK" (id fixo "ok_button"). Esse popup pode
    surgir em QUALQUER momento da execução, então essa função é chamada
    tanto automaticamente (toda vez que a página carrega) quanto
    manualmente em pontos críticos do código, como reforço.
    """
    try:
        botao_ok = page.locator("#ok_button")
        if botao_ok.count() > 0 and botao_ok.first.is_visible():
            botao_ok.first.click(timeout=2000)
            print("  (popup 'chamados em andamento' fechado automaticamente)")
    except Exception:
        pass  # se der qualquer erro nessa checagem, simplesmente ignora e segue


def clicar_se_existir(alvo, localizador, descricao: str) -> bool:
    """
    Clica no elemento se ele existir na tela. Se não existir, assume que
    essa etapa já foi feita por outra pessoa antes do robô chegar nesse
    chamado (ou não se aplica a esse caso específico) e simplesmente pula,
    sem dar erro. Retorna True se clicou, False se pulou.

    Usa force=True no clique: os botões de ação do ServiceNow podem ficar
    escondidos dentro de um menu "overflow" (tipo "...") quando a tela não
    tem espaço pra mostrar todos de uma vez — force ignora a checagem de
    visibilidade e dispara o clique direto no elemento mesmo assim.
    """
    if localizador.count() == 0:
        print(f"  (pulando \"{descricao}\": não encontrado na tela — provavelmente já foi feito)")
        return False
    localizador.first.click(timeout=5000, force=True)
    alvo.wait_for_load_state("networkidle")
    return True


def clicar_esperando(alvo, localizador, descricao: str, segundos: int = 8) -> bool:
    """
    Igual ao clicar_se_existir, mas ESPERA o elemento aparecer por até X
    segundos antes de desistir. Útil pra botões/popups que só surgem alguns
    instantes depois da ação anterior (caso do "Não aplicar material", que
    aparece um pouco depois do "Start Work"). Se mesmo após a espera o
    elemento não aparecer, assume que essa etapa não se aplica e pula sem
    dar erro. Retorna True se clicou, False se pulou.
    """
    try:
        localizador.first.wait_for(state="attached", timeout=segundos * 1000)
    except Exception:
        print(f"  (pulando \"{descricao}\": não apareceu após {segundos}s — provavelmente não se aplica)")
        return False
    localizador.first.click(timeout=5000, force=True)
    alvo.wait_for_load_state("networkidle")
    return True


def validar_tecnico_assignado(alvo, tecnico_planilha: str) -> None:
    """
    Confere se o técnico atribuído no ServiceNow (campo "Atribuído a" /
    "sys_display.wm_task.assigned_to") corresponde ao técnico da planilha
    (coluna F). Usa "contém" (não "começa com"), porque o campo do
    ServiceNow pode ter números ou outras informações antes do nome.
    Compara só pelo primeiro nome, pra tolerar pequenas diferenças de
    digitação no resto do nome. Se não bater, levanta um erro — isso faz
    esse chamado ser pulado e registrado como falha no relatório.
    """
    if not tecnico_planilha:
        return  # sem nome na planilha pra comparar, não tem o que validar

    primeiro_nome = tecnico_planilha.strip().split()[0]
    valor_servicenow = alvo.input_value("input[id='sys_display.wm_task.assigned_to']")

    if primeiro_nome.lower() not in valor_servicenow.lower():
        raise RuntimeError(
            f"Técnico não confere: esperava encontrar '{primeiro_nome}' no campo "
            f"'Atribuído a', mas o valor lá é '{valor_servicenow.strip()}'."
        )


def preencher_etapas_man(page: Page, alvo, inicio_real: str, termino_real: str,
                          validador: str, matricula: str, tecnico: str) -> None:
    """
    Sequência de cliques dentro do chamado MAN, depois do reload:
    Aceitar -> Start Work -> Não aplicar material -> aba "Real" ->
    preencher Início/Término real -> Encerrar Concluído.

    As 3 primeiras etapas são tratadas como opcionais: se o chamado já
    estiver numa etapa mais avançada (porque outra pessoa já clicou em
    "Aceitar", por exemplo), o botão correspondente não vai existir na
    tela — nesse caso o robô simplesmente pula essa etapa e continua a
    partir de onde o chamado realmente está. Só desiste do chamado (pulando
    pro próximo) se NENHUMA das três existir.
    """
    # Confere o técnico ANTES de qualquer clique — evita aceitar/processar
    # um chamado que não é desse técnico.
    validar_tecnico_assignado(alvo, tecnico)

    # 🔧 AJUSTE AQUI (1/5 a 3/5) — Aceitar / Start Work / Não aplicar material
    # ✅ Confirmado pelo HTML: o botão "Aceitar" tem id fixo "accept" — mais
    # confiável que buscar pelo texto.
    loc_aceitar = alvo.locator("#accept")
    loc_start_work = alvo.get_by_text("Start Work", exact=True)
    # ✅ Confirmado pelo HTML: o texto real do botão é "Não Aplicar Material"
    # (com A e M maiúsculos), não "Não aplicar material". Por isso usamos um
    # padrão que ignora maiúsculas/minúsculas, mas ancorado em "^Não...$"
    # pra NÃO confundir com o botão vizinho "Aplicar Material" (sem o "Não").
    loc_nao_aplicar = alvo.get_by_text(re.compile(r"^Não Aplicar Material$", re.IGNORECASE))

    if loc_aceitar.count() == 0 and loc_start_work.count() == 0 and loc_nao_aplicar.count() == 0:
        raise RuntimeError(
            "Nenhuma das etapas iniciais (Aceitar / Start Work / Não aplicar "
            "material) foi encontrada — pulando esse chamado."
        )

    clicar_se_existir(alvo, loc_aceitar, "Aceitar")
    clicar_se_existir(alvo, loc_start_work, "Start Work")
    # "Não aplicar material" costuma aparecer só um pouco depois do "Start
    # Work" — por isso esperamos até alguns segundos por ele, em vez de
    # checar instantaneamente.
    clicar_esperando(alvo, loc_nao_aplicar, "Não aplicar material", segundos=8)

    # ✅ Confirmado: a aba "Real" é um <span class="tab_caption_text">Real</span>.
    alvo.locator("span.tab_caption_text:text-is('Real')").first.click(force=True)
    alvo.wait_for_load_state("networkidle")

    # ✅ Confirmado: os campos são os inputs com id "wm_task.work_start" e
    # "wm_task.work_end" (nomes técnicos do ServiceNow para Início/Término
    # real do trabalho).
    # Campos de data/hora do ServiceNow geralmente só "confirmam" o valor
    # depois que o campo perde o foco — por isso aperta Tab depois de cada
    # preenchimento, simulando o que aconteceria se você preenchesse e
    # clicasse em outro campo.
    alvo.fill("input[id='wm_task.work_start']", inicio_real)
    alvo.press("input[id='wm_task.work_start']", "Tab")

    alvo.fill("input[id='wm_task.work_end']", termino_real)
    alvo.press("input[id='wm_task.work_end']", "Tab")

    # --- Aba "Informações de Fechamento" ---
    # ✅ Confirmado: a aba é um <span class="tab_caption_text">Informações de Fechamento</span>.
    alvo.locator("span.tab_caption_text:text-is('Informações de Fechamento')").first.click(force=True)
    alvo.wait_for_load_state("networkidle")

    # ✅ Confirmado pelo HTML: o código de fechamento é um <select> com id
    # "wm_task.u_fsm_close_code". select_option escolhe a opção "Outros".
    alvo.select_option("select[id='wm_task.u_fsm_close_code']", "Outros")

    # ✅ Confirmado pelo HTML: campos de validador e matrícula.
    alvo.fill("input[id='wm_task.u_validated_by']", validador)
    alvo.fill("input[id='wm_task.u_validated_id_by']", matricula)

    fechar_popup_chamados_pausados(page)  # reforço manual, antes do fechamento

    # ✅ Confirmado pelo HTML: o botão tem id fixo "close_complete" (e o
    # texto real é "Encerrado Concluído", não "Encerrar Concluído"). Usar o
    # id é mais confiável que o texto — não quebra por maiúscula/acento.
    alvo.locator("#close_complete").click(force=True)
    alvo.wait_for_load_state("networkidle")

    print(f"  Início real: {inicio_real} | Término real: {termino_real} preenchidos e encerrado.")

    # >>> PRÓXIMAS ETAPAS SERÃO ADICIONADAS AQUI (você ainda vai me dizer quais) <<<


# ============================================================================
# LOGIN MANUAL (modo --login)
# ============================================================================
def fazer_login_manual(context: BrowserContext) -> None:
    page = context.new_page()
    page.goto(EXCEL_ONLINE_URL)
    print("\n>>> Faça login manualmente na Microsoft nesta janela (Excel Online).")

    page2 = context.new_page()
    page2.goto(SERVICENOW_URL)
    print(">>> Faça login manualmente no ServiceNow nesta outra aba.")

    input("\nDepois de logar nos dois, volte aqui e aperte Enter para salvar a sessão...")
    print("Sessão salva em:", USER_DATA_DIR)


# ============================================================================
# EXECUÇÃO PRINCIPAL
# ============================================================================
def main() -> None:
    modo_login = "--login" in sys.argv

    with sync_playwright() as p:
        # launch_persistent_context salva os cookies/sessão na pasta USER_DATA_DIR,
        # funcionando como um perfil de navegador normal entre execuções.
        # channel="msedge" usa o Microsoft Edge instalado no sistema (o padrão
        # da empresa), em vez do Chromium genérico.
        context = p.chromium.launch_persistent_context(
            user_data_dir=str(USER_DATA_DIR),
            channel="msedge",
            headless=False,  # mantenha False enquanto ajusta os seletores
            accept_downloads=True,
        )

        if modo_login:
            fazer_login_manual(context)
            context.close()
            return

        page = context.pages[0] if context.pages else context.new_page()

        # Aceita automaticamente qualquer diálogo NATIVO do navegador que
        # apareça (precaução — na prática, o popup real do ServiceNow é
        # feito em HTML normal, tratado por fechar_popup_chamados_pausados).
        page.on("dialog", lambda dialog: dialog.accept())

        # Roda essa checagem toda vez que a página terminar de carregar algo
        # (cobre a maioria dos casos, já que os botões do ServiceNow geralmente
        # recarregam a página). Também é chamada manualmente em pontos
        # críticos dentro de preencher_etapas_man, como reforço.
        page.on("load", lambda *_args: fechar_popup_chamados_pausados(page))

        if EXCEL_LOCAL_PATH and Path(EXCEL_LOCAL_PATH).exists():
            print(f"Lendo planilha local em: {EXCEL_LOCAL_PATH}")
            caminho_planilha = Path(EXCEL_LOCAL_PATH)
        else:
            if EXCEL_LOCAL_PATH:
                print(
                    f"⚠️  EXCEL_LOCAL_PATH foi configurado ('{EXCEL_LOCAL_PATH}') mas "
                    "o arquivo não foi encontrado nesse caminho. Caindo para o "
                    "método via Excel Online."
                )
            print("Baixando a planilha do Excel Online...")
            caminho_planilha = baixar_planilha_excel_online(page)

        itens = ler_itens_da_planilha(caminho_planilha)
        print(f"{len(itens)} chamado(s) encontrado(s) na planilha.")

        # --------------------------------------------------------------
        # Preparar a escrita do "Status" de volta na planilha. Só faz
        # sentido escrever se estivermos usando o arquivo local sincronizado
        # (EXCEL_LOCAL_PATH) — escrever na cópia baixada via Excel Online
        # não teria efeito nenhum na planilha real que a equipe usa.
        # --------------------------------------------------------------
        import openpyxl
        permite_escrever_status = bool(EXCEL_LOCAL_PATH) and caminho_planilha == Path(EXCEL_LOCAL_PATH)

        if permite_escrever_status:
            wb_escrita = openpyxl.load_workbook(caminho_planilha)  # sem data_only, preserva fórmulas
            ws_escrita = wb_escrita.active
            coluna_status = obter_ou_criar_coluna_status(ws_escrita)
            wb_escrita.save(caminho_planilha)
            print(f"Status será escrito na planilha local (coluna {coluna_status}).")
        else:
            print(
                "⚠️  Aviso: EXCEL_LOCAL_PATH não está configurado (ou não é o arquivo "
                "em uso agora) — o Status NÃO será escrito de volta na planilha, já "
                "que a cópia baixada do Excel Online é só temporária."
            )

        print("Acessando o ServiceNow...")
        page.goto(SERVICENOW_URL)
        page.wait_for_load_state("networkidle")

        sucesso = []
        falha = []

        for item in itens:
            sgm = item["sgm"]
            try:
                alvo = buscar_chamado_e_clicar_man(page, sgm)
                preencher_etapas_man(
                    page, alvo,
                    item["inicio_real"], item["termino_real"],
                    item["validador"], item["matricula"],
                    item["tecnico"],
                )
                sucesso.append(sgm)
                if permite_escrever_status:
                    ws_escrita.cell(row=item["linha"], column=coluna_status, value="Concluído")
                    wb_escrita.save(caminho_planilha)
            except Exception as erro:
                print(f"Erro ao processar o chamado {sgm}: {erro}")
                falha.append((sgm, str(erro)))
                if permite_escrever_status:
                    ws_escrita.cell(row=item["linha"], column=coluna_status, value=str(erro))
                    wb_escrita.save(caminho_planilha)
                # Screenshot automática pra facilitar o debug depois, sem
                # precisar reproduzir o erro na hora.
                try:
                    caminho_print = ERROS_DIR / f"erro_{sgm}.png"
                    page.screenshot(path=str(caminho_print))
                    print(f"  Screenshot do erro salva em: {caminho_print}")
                except Exception:
                    pass  # se nem o screenshot funcionar, segue o processamento

            time.sleep(1)  # pequena pausa entre chamados, evita sobrecarregar o ServiceNow

        # --------------------------------------------------------------
        # Salva o resultado em arquivo (não só no terminal), pra você
        # poder revisar depois quais SGMs falharam exatamente.
        # --------------------------------------------------------------
        agora = datetime.now().strftime("%Y%m%d_%H%M%S")

        caminho_relatorio = RESULTADOS_DIR / f"relatorio_{agora}.csv"
        with open(caminho_relatorio, "w", newline="", encoding="utf-8-sig") as f:
            escritor = csv.writer(f)
            escritor.writerow(["SGM", "Status", "Motivo"])
            for sgm in sucesso:
                escritor.writerow([sgm, "Sucesso", ""])
            for sgm, motivo in falha:
                escritor.writerow([sgm, "Falha", motivo])

        caminho_falhas = RESULTADOS_DIR / f"sgms_com_falha_{agora}.txt"
        with open(caminho_falhas, "w", encoding="utf-8") as f:
            for sgm, _ in falha:
                f.write(sgm + "\n")

        print("\n--- Resumo da execução ---")
        print(f"Sucesso: {len(sucesso)}/{len(itens)}")
        if falha:
            print(f"Falha: {len(falha)}/{len(itens)}")
            for sgm, motivo in falha:
                print(f"  - {sgm}: {motivo}")
        print(f"\nRelatório completo salvo em: {caminho_relatorio}")
        print(f"Lista só dos SGMs com falha salva em: {caminho_falhas}")

        context.close()


if __name__ == "__main__":
    main()
